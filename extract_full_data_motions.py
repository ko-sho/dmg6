import openpyxl
import os
import json
from typing import Any

# 入力ファイル
FULL_DATA_FILE = "full_data_1.0.11.xlsx"
# 出力先ディレクトリ
OUT_DIR = os.path.join("src", "data", "fullDataWeapons")

# 武器シート名リスト（full_dataのシート名）
WEAPON_SHEETS = [
    'wp00GS', 'wp01Sns', 'wp02DB', 'wp03LS', 'wp04Ham', 'wp05HH',
    'wp06Lan', 'wp07GL', 'wp08SA', 'wp09CB', 'wp10IG', 'wp11Bow',
    # 'wp12HBG', 'wp13LBG',  # ボウガン系は一旦除外
]

# 2行目のカラム名を取得して型に合わせてdictを作る
def extract_full_data_motions():
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(FULL_DATA_FILE, data_only=True)
    for ws_name in WEAPON_SHEETS:
        if ws_name not in wb.sheetnames:
            continue
        ws = wb[ws_name]
        # 2行目がカラム名
        headers = [cell.value for cell in ws[2]]
        # 4行目以降がデータ
        motions = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not any(row):
                continue
            d = {h: (None if v == "NULL" else v) for h, v in zip(headers, row)}
            # RSIDがNoneまたは空文字の場合のみ除外（0や'0'は許容）
            if d.get("RSID") is None or d.get("RSID") == "":
                continue
            # 必須カラムがすべて存在しない場合はスキップ
            if not (d.get("colName") or d.get("RSName")):
                continue
            if d.get("Attack") is None:
                continue
            if d.get("StatusAttrRate") is None:
                continue
            if d.get("IsSkillHien") is None:
                continue
            d["name"] = d.get("colName") or d.get("RSName") or ""
            # A列（1列目）の英語名をenNameに優先して格納
            a_col_en = row[0] if len(row) > 0 and row[0] else None
            d["enName"] = a_col_en or d.get("RSName") or d.get("colName") or ""
            # key: name + '-' + RSID
            d["key"] = f"{d['name']}-{d['RSID']}"
            motions.append(d)
        # RSID順にソート
        motions.sort(key=lambda d: int(d["RSID"]))
        motions.sort(key=lambda d: int(d["GroupIndex"]) if d.get("GroupIndex") == 0 else 99)  # GroupIndexがない場合は99に設定
        # ファイル名例: wp00GS.json
        out_path = os.path.join(OUT_DIR, f"{ws_name}.json")
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(motions, f, ensure_ascii=False, indent=2)
        print(f"[INFO] {out_path} written ({len(motions)} motions)")

if __name__ == "__main__":
    extract_full_data_motions()
